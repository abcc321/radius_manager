from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from common import get_db, NetworkUser, Apartment, Plan, hash_password, model_to_dict
from modules.audit_log.utils import log_audit

router = APIRouter(prefix="/network_users", tags=["网络用户管理"])


class NetworkUserCreate(BaseModel):
    apartment_id: int
    username: str
    password: str
    name: Optional[str] = None
    phone: Optional[str] = None
    room: Optional[str] = None
    plan_id: Optional[int] = None
    activate_date: Optional[str] = None
    expire_date: Optional[str] = None


class NetworkUserUpdate(BaseModel):
    apartment_id: Optional[int] = None
    name: Optional[str] = None
    phone: Optional[str] = None
    room: Optional[str] = None
    plan_id: Optional[int] = None
    activate_date: Optional[str] = None
    expire_date: Optional[str] = None


class NetworkUserPasswordUpdate(BaseModel):
    password: str


@router.get("/")
async def get_network_users(
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    status: Optional[str] = Query(None, description="状态筛选"),
    apartment_id: Optional[int] = Query(None, description="公寓ID筛选"),
    plan_id: Optional[int] = Query(None, description="套餐ID筛选"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """获取网络用户列表"""
    query = db.query(NetworkUser)

    if keyword:
        query = query.filter(
            or_(
                NetworkUser.username.contains(keyword),
                NetworkUser.name.contains(keyword),
                NetworkUser.phone.contains(keyword),
                NetworkUser.room.contains(keyword)
            )
        )

    if status:
        today = datetime.now().strftime("%Y-%m-%d")
        one_month_later = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")

        if status == "active":
            query = query.filter(
                and_(
                    NetworkUser.expire_date.isnot(None),
                    NetworkUser.expire_date > one_month_later
                )
            )
        elif status == "expiring":
            query = query.filter(
                and_(
                    NetworkUser.expire_date.isnot(None),
                    NetworkUser.expire_date > today,
                    NetworkUser.expire_date <= one_month_later
                )
            )
        elif status == "expired":
            query = query.filter(
                and_(
                    NetworkUser.expire_date.isnot(None),
                    NetworkUser.expire_date <= today
                )
            )
        elif status == "inactive":
            query = query.filter(
                or_(
                    NetworkUser.expire_date.is_(None),
                    NetworkUser.expire_date == ""
                )
            )

    if apartment_id:
        query = query.filter(NetworkUser.apartment_id == apartment_id)

    if plan_id:
        query = query.filter(NetworkUser.plan_id == plan_id)

    total = query.count()
    users = query.order_by(NetworkUser.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    result = []
    for u in users:
        user_dict = model_to_dict(u)

        apt = db.query(Apartment).filter(Apartment.id == u.apartment_id).first()
        user_dict["apartment_name"] = apt.name if apt else None

        if u.plan_id:
            plan = db.query(Plan).filter(Plan.id == u.plan_id).first()
            user_dict["plan_name"] = plan.name if plan else None
            user_dict["plan_price"] = plan.price if plan else None
        else:
            user_dict["plan_name"] = None
            user_dict["plan_price"] = None

        result.append(user_dict)

    return {
        "code": 200,
        "data": result,
        "total": total,
        "page": page,
        "page_size": page_size
    }


@router.get("/all")
async def get_all_network_users(
    status: Optional[str] = Query("active", description="状态筛选"),
    db: Session = Depends(get_db)
):
    """获取所有网络用户（下拉框用）"""
    query = db.query(NetworkUser)

    if status:
        query = query.filter(NetworkUser.status == status)

    users = query.all()

    result = []
    for u in users:
        user_dict = model_to_dict(u)

        apt = db.query(Apartment).filter(Apartment.id == u.apartment_id).first()
        user_dict["apartment_name"] = apt.name if apt else None

        if u.plan_id:
            plan = db.query(Plan).filter(Plan.id == u.plan_id).first()
            user_dict["plan_name"] = plan.name if plan else None
        else:
            user_dict["plan_name"] = None

        result.append(user_dict)

    return {
        "code": 200,
        "data": result
    }


@router.get("/export-users")
async def export_network_users(
    status: Optional[str] = Query(None, description="状态筛选"),
    apartment_id: Optional[int] = Query(None, description="公寓ID筛选"),
    db: Session = Depends(get_db)
):
    """导出网络用户"""
    query = db.query(NetworkUser)

    if status and status != "":
        today = datetime.now().strftime("%Y-%m-%d")
        one_month_later = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")

        if status == "active":
            query = query.filter(
                and_(
                    NetworkUser.expire_date.isnot(None),
                    NetworkUser.expire_date > one_month_later
                )
            )
        elif status == "expiring":
            query = query.filter(
                and_(
                    NetworkUser.expire_date.isnot(None),
                    NetworkUser.expire_date > today,
                    NetworkUser.expire_date <= one_month_later
                )
            )
        elif status == "expired":
            query = query.filter(
                and_(
                    NetworkUser.expire_date.isnot(None),
                    NetworkUser.expire_date <= today
                )
            )
        elif status == "inactive":
            query = query.filter(
                or_(
                    NetworkUser.expire_date.is_(None),
                    NetworkUser.expire_date == ""
                )
            )

    if apartment_id:
        query = query.filter(NetworkUser.apartment_id == apartment_id)

    users = query.order_by(NetworkUser.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "网络用户"

    headers = ["上网账号", "密码", "姓名", "手机号", "房间号", "公寓", "套餐", "状态", "开通日期", "到期日期"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for u in users:
        apt = db.query(Apartment).filter(Apartment.id == u.apartment_id).first()
        apt_name = apt.name if apt else ""

        plan_name = ""
        plan_price = ""
        if u.plan_id:
            plan = db.query(Plan).filter(Plan.id == u.plan_id).first()
            if plan:
                plan_name = plan.name
                plan_price = plan.price

        if not u.expire_date:
            status_text = "未开通"
        else:
            expire_date_obj = datetime.strptime(u.expire_date, "%Y-%m-%d")
            now = datetime.now()
            one_month_later = now + timedelta(days=30)

            if expire_date_obj <= now:
                status_text = "已过期"
            elif expire_date_obj <= one_month_later:
                status_text = "即将到期"
            else:
                status_text = "已开通"

        row = [
            u.username,
            u.password or "",
            u.name or "",
            u.phone or "",
            u.room or "",
            apt_name,
            f"{plan_name} ({plan_price})" if plan_name else "",
            status_text,
            u.activate_date or "",
            u.expire_date or ""
        ]
        ws.append(row)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"network_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.post("/import-users")
async def import_network_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """批量导入网络用户"""
    print(f"[IMPORT] Received file: {file.filename}")

    if not file.filename.endswith(('.xlsx', '.xls')):
        print(f"[IMPORT] Invalid file format: {file.filename}")
        raise HTTPException(status_code=400, detail="只支持Excel文件")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        print(f"[IMPORT] Excel headers: {headers}")
        required_headers = ["上网账号", "密码", "姓名", "手机号", "房间号", "公寓", "套餐", "开通日期", "到期日期"]

        for required in required_headers:
            if required not in headers:
                print(f"[IMPORT] Missing required header: {required}")
                raise HTTPException(status_code=400, detail=f"缺少必需列: {required}")

        success_count = 0
        error_count = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                username = row[headers.index("上网账号")]
                password = row[headers.index("密码")]
                name = row[headers.index("姓名")]
                phone = str(row[headers.index("手机号")]) if row[headers.index("手机号")] else ""
                room = row[headers.index("房间号")]
                apartment_name = row[headers.index("公寓")]
                plan_name = row[headers.index("套餐")]
                activate_date = row[headers.index("开通日期")]
                expire_date = row[headers.index("到期日期")]

                if not username or not password:
                    errors.append(f"第{row_idx}行: 上网账号和密码不能为空")
                    error_count += 1
                    continue

                apt = db.query(Apartment).filter(Apartment.name == apartment_name).first()
                if not apt:
                    errors.append(f"第{row_idx}行: 公寓 {apartment_name} 不存在")
                    error_count += 1
                    continue

                existing_user = db.query(NetworkUser).filter(
                    NetworkUser.apartment_id == apt.id,
                    NetworkUser.username == username
                ).first()
                if existing_user:
                    errors.append(f"第{row_idx}行: 公寓 {apartment_name} 内上网账号 {username} 已存在")
                    error_count += 1
                    continue

                plan_id = None
                if plan_name:
                    plan_name_clean = plan_name.split(" (")[0] if " (" in plan_name else plan_name
                    plan = db.query(Plan).filter(Plan.name == plan_name_clean).first()
                    if plan:
                        plan_id = plan.id

                activate_date_str = activate_date.strftime("%Y-%m-%d") if isinstance(activate_date, datetime) else str(activate_date) if activate_date else None
                expire_date_str = expire_date.strftime("%Y-%m-%d") if isinstance(expire_date, datetime) else str(expire_date) if expire_date else None

                user = NetworkUser(
                    apartment_id=apt.id,
                    username=username,
                    password=password,
                    name=name,
                    phone=phone,
                    room=room,
                    plan_id=plan_id,
                    status="active",
                    activate_date=activate_date_str,
                    expire_date=expire_date_str
                )
                db.add(user)
                success_count += 1

            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
                error_count += 1

        db.commit()

        return {
            "code": 200,
            "message": f"导入完成",
            "data": {
                "success_count": success_count,
                "error_count": error_count,
                "errors": errors[:10]
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.get("/{user_id}")
async def get_network_user(user_id: int, db: Session = Depends(get_db)):
    """获取网络用户详情"""
    user = db.query(NetworkUser).filter(NetworkUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    user_dict = model_to_dict(user)

    apt = db.query(Apartment).filter(Apartment.id == user.apartment_id).first()
    user_dict["apartment_name"] = model_to_dict(apt) if apt else None

    if user.plan_id:
        plan = db.query(Plan).filter(Plan.id == user.plan_id).first()
        if plan:
            user_dict["plan"] = model_to_dict(plan)
            user_dict["plan_price"] = plan.price
            user_dict["plan_upload_speed"] = plan.upload_speed
            user_dict["plan_download_speed"] = plan.download_speed
        else:
            user_dict["plan"] = None
            user_dict["plan_price"] = None
    else:
        user_dict["plan"] = None
        user_dict["plan_price"] = None

    return {"code": 200, "data": user_dict}


@router.post("/")
async def create_network_user(user_data: NetworkUserCreate, request: Request, db: Session = Depends(get_db)):
    """创建网络用户"""
    existing_user = db.query(NetworkUser).filter(
        NetworkUser.apartment_id == user_data.apartment_id,
        NetworkUser.username == user_data.username
    ).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="该公寓内上网账号已存在")

    apt = db.query(Apartment).filter(Apartment.id == user_data.apartment_id).first()
    if not apt:
        raise HTTPException(status_code=400, detail="公寓不存在")

    if user_data.plan_id:
        plan = db.query(Plan).filter(Plan.id == user_data.plan_id).first()
        if not plan:
            raise HTTPException(status_code=400, detail="套餐不存在")

    user = NetworkUser(
        apartment_id=user_data.apartment_id,
        username=user_data.username,
        password=user_data.password,
        name=user_data.name,
        phone=user_data.phone,
        room=user_data.room,
        plan_id=user_data.plan_id,
        status="active",
        activate_date=user_data.activate_date or datetime.now().strftime("%Y-%m-%d"),
        expire_date=user_data.expire_date
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    # 记录审计日志
    plan_name = None
    if user_data.plan_id:
        plan = db.query(Plan).filter(Plan.id == user_data.plan_id).first()
        if plan:
            plan_name = plan.name

    log_audit(
        db=db,
        operator_id=getattr(request.state, 'operator_id', None),
        operator_name=getattr(request.state, 'operator_name', None),
        module="网络用户管理",
        action="CREATE",
        target_type="NetworkUser",
        target_id=user.id,
        target_name=user.username,
        description=f"创建网络用户：{user.username}，姓名：{user.name or '未填写'}，公寓：{apt.name}",
        new_data=model_to_dict(user),
        ip_address=request.client.host if request.client else None,
        status="success"
    )

    return {
        "code": 200,
        "message": "创建成功",
        "data": model_to_dict(user)
    }


@router.put("/{user_id}")
async def update_network_user(
    user_id: int,
    user_data: NetworkUserUpdate,
    request: Request,
    db: Session = Depends(get_db)
):
    """更新网络用户"""
    user = db.query(NetworkUser).filter(NetworkUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 记录修改前的数据
    old_data = model_to_dict(user)

    if user_data.apartment_id is not None:
        apt = db.query(Apartment).filter(Apartment.id == user_data.apartment_id).first()
        if not apt:
            raise HTTPException(status_code=400, detail="公寓不存在")
        user.apartment_id = user_data.apartment_id

    if user_data.name is not None:
        user.name = user_data.name
    if user_data.phone is not None:
        user.phone = user_data.phone
    if user_data.room is not None:
        user.room = user_data.room
    if user_data.plan_id is not None:
        if user_data.plan_id > 0:
            plan = db.query(Plan).filter(Plan.id == user_data.plan_id).first()
            if not plan:
                raise HTTPException(status_code=400, detail="套餐不存在")
        user.plan_id = user_data.plan_id if user_data.plan_id > 0 else None
    if user_data.activate_date is not None:
        user.activate_date = user_data.activate_date
    if user_data.expire_date is not None:
        user.expire_date = user_data.expire_date

    db.commit()
    db.refresh(user)

    # 记录审计日志
    log_audit(
        db=db,
        operator_id=getattr(request.state, 'operator_id', None),
        operator_name=getattr(request.state, 'operator_name', None),
        module="网络用户管理",
        action="UPDATE",
        target_type="NetworkUser",
        target_id=user.id,
        target_name=user.username,
        description=f"更新网络用户：{user.username}",
        old_data=old_data,
        new_data=model_to_dict(user),
        ip_address=request.client.host if request.client else None,
        status="success"
    )

    return {
        "code": 200,
        "message": "更新成功",
        "data": model_to_dict(user)
    }


@router.post("/{user_id}/deactivate")
async def deactivate_network_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    """停用网络用户"""
    user = db.query(NetworkUser).filter(NetworkUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 记录修改前的数据
    old_data = model_to_dict(user)

    user.activate_date = None
    user.expire_date = None
    user.status = "inactive"

    db.commit()
    db.refresh(user)

    # 记录审计日志
    log_audit(
        db=db,
        operator_id=getattr(request.state, 'operator_id', None),
        operator_name=getattr(request.state, 'operator_name', None),
        module="网络用户管理",
        action="UPDATE",
        target_type="NetworkUser",
        target_id=user.id,
        target_name=user.username,
        description=f"停用网络用户：{user.username}",
        old_data=old_data,
        new_data=model_to_dict(user),
        ip_address=request.client.host if request.client else None,
        status="success"
    )

    return {
        "code": 200,
        "message": "停用成功",
        "data": model_to_dict(user)
    }


@router.put("/{user_id}/password")
async def update_network_user_password(
    user_id: int,
    password_data: NetworkUserPasswordUpdate,
    request: Request,
    db: Session = Depends(get_db)
):
    """修改网络用户密码"""
    user = db.query(NetworkUser).filter(NetworkUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 记录修改前的数据（不包含密码明文）
    old_data = {"id": user.id, "username": user.username, "status": user.status}

    user.password = password_data.password
    db.commit()

    # 记录审计日志
    log_audit(
        db=db,
        operator_id=getattr(request.state, 'operator_id', None),
        operator_name=getattr(request.state, 'operator_name', None),
        module="网络用户管理",
        action="UPDATE",
        target_type="NetworkUser",
        target_id=user.id,
        target_name=user.username,
        description=f"修改网络用户密码：{user.username}",
        old_data=old_data,
        new_data={"id": user.id, "username": user.username, "status": user.status},
        ip_address=request.client.host if request.client else None,
        status="success"
    )

    return {
        "code": 200,
        "message": "密码修改成功"
    }


@router.post("/{user_id}/activate")
async def activate_network_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    """开通网络用户"""
    user = db.query(NetworkUser).filter(NetworkUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 记录修改前的数据
    old_data = model_to_dict(user)

    user.status = "active"
    user.activate_date = datetime.now().strftime("%Y-%m-%d")
    db.commit()

    # 记录审计日志
    log_audit(
        db=db,
        operator_id=getattr(request.state, 'operator_id', None),
        operator_name=getattr(request.state, 'operator_name', None),
        module="网络用户管理",
        action="UPDATE",
        target_type="NetworkUser",
        target_id=user.id,
        target_name=user.username,
        description=f"开通网络用户：{user.username}",
        old_data=old_data,
        new_data=model_to_dict(user),
        ip_address=request.client.host if request.client else None,
        status="success"
    )

    return {
        "code": 200,
        "message": "开通成功",
        "data": {
            "status": user.status,
            "activate_date": user.activate_date
        }
    }


@router.delete("/{user_id}")
async def delete_network_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    """删除网络用户"""
    user = db.query(NetworkUser).filter(NetworkUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    # 记录删除前的数据
    old_data = model_to_dict(user)
    username = user.username

    db.delete(user)
    db.commit()

    # 记录审计日志
    log_audit(
        db=db,
        operator_id=getattr(request.state, 'operator_id', None),
        operator_name=getattr(request.state, 'operator_name', None),
        module="网络用户管理",
        action="DELETE",
        target_type="NetworkUser",
        target_id=user_id,
        target_name=username,
        description=f"删除网络用户：{username}",
        old_data=old_data,
        ip_address=request.client.host if request.client else None,
        status="success"
    )

    return {"code": 200, "message": "删除成功"}

"""
生成账单服务
"""
import os
from datetime import datetime
from typing import List, Dict, Optional
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

from common import Apartment, NetworkUser, Plan, BillRecord


class BillingGenerateService:
    """生成账单服务"""

    def __init__(self, db: Session):
        self.db = db
        self.base_dir = Path("d:/trae_project/radius_manager/backend/downloads/bills")
        self.base_dir.mkdir(parents=True, exist_ok=True)

    def calculate_monthly_fee(
        self,
        activate_date: str,
        expire_date: str,
        plan_price: float,
        target_year: int,
        target_month: int
    ) -> float:
        """计算指定月份的月费"""
        if not activate_date or not expire_date or not plan_price:
            return 0.0

        try:
            activate = datetime.strptime(activate_date, "%Y-%m-%d")
            expire = datetime.strptime(expire_date, "%Y-%m-%d")

            from datetime import timedelta
            month_start = datetime(target_year, target_month, 1)
            if target_month == 12:
                month_end = datetime(target_year + 1, 1, 1) - timedelta(days=1)
            else:
                month_end = datetime(target_year, target_month + 1, 1) - timedelta(days=1)

            if activate > month_end or expire < month_start:
                return 0.0

            is_first_month = activate.year == target_year and activate.month == target_month
            is_last_month = expire.year == target_year and expire.month == target_month

            if is_first_month:
                days_used = month_end.day - activate.day + 1
                return 0.0 if days_used < 26 else float(plan_price)
            elif is_last_month:
                return 0.0 if expire.day < 26 else float(plan_price)
            else:
                return float(plan_price)
        except:
            return 0.0

    def generate_bill(
        self,
        apartment_id: int,
        year: int,
        month: int,
        operator_id: Optional[int] = None,
        operator_name: Optional[str] = None
    ) -> Dict:
        """生成账单"""
        apartment = self.db.query(Apartment).filter(
            Apartment.id == apartment_id,
            Apartment.status != "deleted"
        ).first()

        if not apartment:
            raise ValueError(f"公寓不存在: {apartment_id}")

        users = self.db.query(NetworkUser).filter(
            NetworkUser.apartment_id == apartment_id
        ).all()

        total_accounts = len(users)
        active_accounts = sum(1 for u in users if u.status == "active")
        inactive_accounts = total_accounts - active_accounts

        bill_details = []
        total_amount = 0.0

        for user in users:
            plan_info = None
            if user.plan_id:
                plan = self.db.query(Plan).filter(Plan.id == user.plan_id).first()
                if plan:
                    plan_info = {
                        "name": plan.name,
                        "price": plan.price
                    }

            fee = self.calculate_monthly_fee(
                user.activate_date,
                user.expire_date,
                plan_info["price"] if plan_info else 0,
                year,
                month
            )

            total_amount += fee

            bill_details.append({
                "username": user.username,
                "room": user.room or "",
                "plan_name": plan_info["name"] if plan_info else "",
                "plan_price": float(plan_info["price"]) if plan_info and plan_info["price"] else 0.0,
                "activate_date": user.activate_date or "",
                "expire_date": user.expire_date or "",
                "status": user.status,
                "monthly_fee": fee
            })

        filename = f"bill_{apartment.code}_{year}_{month:02d}.xlsx"
        file_path = self.base_dir / filename

        self._create_excel(file_path, apartment, bill_details, year, month, total_amount)

        bill_record = BillRecord(
            bill_month=f"{year}-{month:02d}",
            apartment_id=apartment_id,
            apartment_name=apartment.name,
            apartment_code=apartment.code,
            total_accounts=total_accounts,
            active_accounts=active_accounts,
            inactive_accounts=inactive_accounts,
            total_amount=f"{total_amount:.2f}",
            file_path=str(file_path),
            file_name=filename,
            file_size=file_path.stat().st_size,
            operator_id=operator_id,
            operator_name=operator_name or "",
            status="completed"
        )

        self.db.add(bill_record)
        self.db.commit()
        self.db.refresh(bill_record)

        return {
            "id": bill_record.id,
            "bill_month": bill_record.bill_month,
            "apartment_name": bill_record.apartment_name,
            "total_accounts": bill_record.total_accounts,
            "total_amount": bill_record.total_amount,
            "file_name": bill_record.file_name,
            "status": bill_record.status,
            "created_at": bill_record.created_at.strftime("%Y-%m-%d %H:%M:%S")
        }

    def _create_excel(
        self,
        file_path: Path,
        apartment: Apartment,
        bill_details: List[Dict],
        year: int,
        month: int,
        total_amount: float
    ):
        """创建Excel文件"""
        wb = Workbook()

        ws = wb.active
        ws.title = "账单汇总"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        summary_data = [
            ["公寓编号", apartment.code],
            ["公寓名称", apartment.name],
            ["账单月份", f"{year}年{month}月"],
            ["总账号数", len(bill_details)],
            ["开通账号", sum(1 for d in bill_details if d["status"] == "active")],
            ["未开通账号", sum(1 for d in bill_details if d["status"] != "active")],
            ["总费用", f"¥{total_amount:.2f}"]
        ]

        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
            ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30

        details_ws = wb.create_sheet(title="账单明细")

        detail_headers = [
            "宽带账号", "房间号", "套餐名称", "月费",
            "开通时间", "到期时间", "状态", "本月费用"
        ]

        details_ws.append(detail_headers)

        for col_idx, _ in enumerate(detail_headers, 1):
            cell = details_ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for detail in bill_details:
            details_ws.append([
                detail["username"],
                detail["room"],
                detail["plan_name"],
                f"¥{detail['plan_price']:.2f}",
                detail["activate_date"],
                detail["expire_date"],
                "已开通" if detail["status"] == "active" else "已停用",
                f"¥{detail['monthly_fee']:.2f}" if detail["monthly_fee"] > 0 else "免费"
            ])

        for row in details_ws.iter_rows(min_row=2, max_row=details_ws.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in details_ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            details_ws.column_dimensions[column_letter].width = min(adjusted_width, 25)

        wb.save(file_path)

    def get_bill_records(
        self,
        apartment_id: Optional[int] = None,
        bill_month: Optional[str] = None,
        page: int = 1,
        page_size: int = 20
    ) -> Dict:
        """获取账单记录列表"""
        query = self.db.query(BillRecord)

        if apartment_id:
            query = query.filter(BillRecord.apartment_id == apartment_id)

        if bill_month:
            query = query.filter(BillRecord.bill_month == bill_month)

        total = query.count()
        records = query.order_by(BillRecord.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "data": [
                {
                    "id": r.id,
                    "bill_month": r.bill_month,
                    "apartment_id": r.apartment_id,
                    "apartment_name": r.apartment_name,
                    "apartment_code": r.apartment_code,
                    "total_accounts": r.total_accounts,
                    "active_accounts": r.active_accounts,
                    "inactive_accounts": r.inactive_accounts,
                    "total_amount": r.total_amount,
                    "file_name": r.file_name,
                    "file_size": r.file_size,
                    "operator_name": r.operator_name,
                    "status": r.status,
                    "created_at": r.created_at.strftime("%Y-%m-%d %H:%M:%S")
                }
                for r in records
            ]
        }

    def get_bill_record(self, record_id: int) -> Optional[Dict]:
        """获取账单记录详情"""
        record = self.db.query(BillRecord).filter(BillRecord.id == record_id).first()

        if not record:
            return None

        return {
            "id": record.id,
            "bill_month": record.bill_month,
            "apartment_id": record.apartment_id,
            "apartment_name": record.apartment_name,
            "apartment_code": record.apartment_code,
            "total_accounts": record.total_accounts,
            "active_accounts": record.active_accounts,
            "inactive_accounts": record.inactive_accounts,
            "total_amount": record.total_amount,
            "file_name": record.file_name,
            "file_path": record.file_path,
            "file_size": record.file_size,
            "operator_name": record.operator_name,
            "status": record.status,
            "created_at": record.created_at.strftime("%Y-%m-%d %H:%M:%S")
        }

    def get_bill_file(self, record_id: int) -> Optional[bytes]:
        """获取账单文件内容"""
        record = self.db.query(BillRecord).filter(BillRecord.id == record_id).first()

        if not record or not record.file_path:
            return None

        file_path = Path(record.file_path)
        if not file_path.exists():
            return None

        with open(file_path, 'rb') as f:
            return f.read()

    def get_bill_details(self, record_id: int) -> Optional[Dict]:
        """获取账单明细"""
        record = self.db.query(BillRecord).filter(BillRecord.id == record_id).first()

        if not record:
            return None

        year, month = map(int, record.bill_month.split('-'))

        users = self.db.query(NetworkUser).filter(
            NetworkUser.apartment_id == record.apartment_id
        ).all()

        bill_details = []
        for user in users:
            plan_info = None
            if user.plan_id:
                plan = self.db.query(Plan).filter(Plan.id == user.plan_id).first()
                if plan:
                    plan_info = {
                        "name": plan.name,
                        "price": plan.price
                    }

            fee = self.calculate_monthly_fee(
                user.activate_date,
                user.expire_date,
                plan_info["price"] if plan_info else 0,
                year,
                month
            )

            bill_details.append({
                "username": user.username,
                "room": user.room or "",
                "plan_name": plan_info["name"] if plan_info else "",
                "plan_price": float(plan_info["price"]) if plan_info and plan_info["price"] else 0.0,
                "activate_date": user.activate_date or "",
                "expire_date": user.expire_date or "",
                "status": user.status,
                "monthly_fee": fee
            })

        return {
            "id": record.id,
            "bill_month": record.bill_month,
            "apartment_id": record.apartment_id,
            "apartment_name": record.apartment_name,
            "apartment_code": record.apartment_code,
            "total_accounts": record.total_accounts,
            "active_accounts": record.active_accounts,
            "inactive_accounts": record.inactive_accounts,
            "total_amount": record.total_amount,
            "file_name": record.file_name,
            "file_path": record.file_path,
            "file_size": record.file_size,
            "operator_name": record.operator_name,
            "status": record.status,
            "created_at": record.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "bill_details": bill_details
        }

    def delete_bill(self, record_id: int) -> bool:
        """删除账单（删除数据库记录和文件）"""
        record = self.db.query(BillRecord).filter(BillRecord.id == record_id).first()

        if not record:
            return False

        if record.file_path:
            file_path = Path(record.file_path)
            if file_path.exists():
                try:
                    file_path.unlink()
                except Exception:
                    pass

        self.db.delete(record)
        self.db.commit()

        return True

    def get_all_apartments(self) -> List[Dict]:
        """获取所有公寓列表"""
        apartments = self.db.query(Apartment).filter(Apartment.status != "deleted").all()
        return [
            {
                "id": apt.id,
                "code": apt.code,
                "name": apt.name
            }
            for apt in apartments
        ]
